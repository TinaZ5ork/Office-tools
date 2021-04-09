from docxtpl import DocxTemplate
from openpyxl import load_workbook
import datetime


def replace(obj):
    if obj is None:
        obj = ''
        return obj


today = datetime.date.today()
# 加载要填入的数据
wb = load_workbook(r"new file.xlsx")  # 需要填入Word的Excel工作簿的地址
ws = wb['CEE']  # 工作簿中表格的名称
tpl = DocxTemplate('fm.docx')

contexts = []
Today = today
Totalcase = ws["C2"].value  # 字母代表表格中对应的列，顺序和列名一定要对应上
newcase = ws["D29"].value
Totaldeath = ws["E29"].value
newdeath = ws["F29"].value

P1 = ws["C2"].value
P2 = ws["D2"].value
P3 = ws["E2"].value
P4 = ws["F2"].value

U1 = ws["C3"].value
U2 = ws["D3"].value
U3 = ws["E3"].value
U4 = ws["F3"].value

R1 = ws["C5"].value
R2 = ws["D5"].value
R3 = ws["E5"].value
R4 = ws["F5"].value

C1 = ws["C4"].value
C2 = ws["D4"].value
C3 = ws["E4"].value
C4 = ws["F4"].value

H1 = ws["C7"].value
H2 = ws["D7"].value
H3 = ws["E7"].value
H4 = ws["F7"].value

A1 = ws["C9"].value
A2 = ws["D9"].value
A3 = ws["E9"].value
A4 = ws["F9"].value

S1 = ws["C8"].value
S2 = ws["D8"].value
S3 = ws["E8"].value
S4 = ws["F8"].value

B1 = ws["C11"].value
B2 = ws["D11"].value
B3 = ws["E11"].value
B4 = ws["F11"].value

M1 = ws["C12"].value
M2 = ws["D12"].value
M3 = ws["E12"].value
M4 = ws["F12"].value

SK1 = ws["C10"].value
SK2 = ws["D10"].value
SK3 = ws["E10"].value
SK4 = ws["F10"].value

ws = wb['Nordic']
Nordic1 = ws["C7"].value
Nordic2 = ws["E7"].value
Sweden = ws["D2"].value
context = {"Totalcase": Totalcase, "newcase": newcase, "Totaldeath": Totaldeath, "newdeath": newdeath,
           "P1": P1, "P2": P2, "P3": P3, "P4": P4,
           "U1": U1, "U2": U2, "U3": U3, "U4": U4,
           "R1": R1, "R2": R2, "R3": R3, "R4": R4,
           "C1": C1, "C2": C2, "C3": C3, "C4": C4,
           "H1": H1, "H2": H2, "H3": H3, "H4": H4,
           "A1": A1, "A2": A2, "A3": A3, "A4": A4,
           "S1": S1, "S2": S2, "S3": S3, "S4": S4,
           "B1": B1, "B2": B2, "B3": S3, "B4": S4,
           "M1": M1, "M2": M2, "M3": M3, "M4": M4,
           "SK1": SK1, "SK2": SK2, "SK3": SK3, "SK4": SK4,
           "Today": Today, "Nordic1": Nordic1, "Nordic2": Nordic2, "Sweden": Sweden}  # 变量名称与Word文档中的占位符要一一对应
contexts.append(context)

for context in contexts:
    print(context)
    tpl = DocxTemplate(r"fm.docx")
    tpl.render(context)
    tpl.save(f'C:/Users/Tina/PycharmProjects/pythonProject1/dist/{Today}.docx')
